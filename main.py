from openpyxl import load_workbook
from openpyxl import Workbook
import urllib3
import configparser
import os
import re


def prepare_config(filename):
    if not os.path.exists(filename):
        with open(filename, 'w+') as f:
            f.write('[global_settings]\n'
                    '# symbols to replace in filenames\n'
                    'url_replace = +, ,/,%%,*\n'
                    '# 64k, increase if needed\n'
                    'data_chunks = 65536\n'
                    '# photos limit, put -1 for no limit\n'
                    'max_photos = -1\n'
                    '# excel starting row\n'
                    'first_row = 2\n'
                    '# for entire file set to -1\n'
                    'last_row = -1\n'
                    'default_file_format = .jpg\n'
                    '# Set True to skip links ending by /\n'
                    '# sipping http://exmlpe.com/photos/\n'
                    '# not skipping http://exmlpe.com/photos/photo.jpg'
                    'skip_trailing_slashes = True\n'
                    '\n'
                    '[paths]\n'
                    '# file must contain artikul in first column and photos in others\n'
                    'input_file = input.xlsx\n'
                    'output_folder = output\n'
                    'output_file = output.xlsx\n'
                    '# enable to put each row to separate folder\n'
                    'create_art_folders = True\n'
                    'site_path = /_shop/\n'
                    '# disable to use backslashes (\\) in paths\n'
                    'nix_paths = True'
                    )
        config_obj = None
    else:
        config_obj = configparser.ConfigParser()
        config_obj.read(filename)
    return config_obj


def prepare_for_url(text):
    text = text.lower()
    for symbol in url_replace:
        text = text.replace(symbol, "-")
    return text


def read_excel(excel_file):
    out_struct = {}
    wb = load_workbook(excel_file)
    ws = wb.active

    regex_start = re.compile(r'^\s\s*')
    regex_end = re.compile(r'\s\s*$')

    for row in ws.iter_rows(min_row=first_row, max_row=last_row, values_only=True):
        row_arr = []
        for cell in row[1:]:
            if cell:
                cell = regex_start.sub('', cell)
                cell = regex_end.sub('', cell)
                if cell[-1::1] == '/' and skip_trailing_slashes:
                    pass
                else:
                    row_arr.append(cell)
        if row[0]:
            out_struct[row[0]] = row_arr

    return out_struct


def write_excel(input_struct):
    wb = Workbook()
    ws = wb.active
    current_row = 0
    for key in input_struct.keys():
        current_row += 1
        current_col = 1
        ws.cell(row=current_row, column=current_col).value = key
        for photo in input_struct[key]:
            current_col += 1
            ws.cell(row=current_row, column=current_col).value = photo
    wb.save(output_file)


def download_art_photo(input_link_struct, path='.', folders=False):
    http = urllib3.PoolManager()
    output_files_struct = {}
    empty_folders_counter = 0
    for art, photos_arr in input_link_struct.items():
        if not photos_arr:
            print('No photo for', art, '!')
            empty_folders_counter += 1
            continue
        prepared_art = prepare_for_url(str(art))
        output_files_struct[art] = []
        print('Downloading', prepared_art)
        files_iterator = 0

        if create_art_folders:
            photo_dir_path = prepared_art
        else:
            photo_dir_path = None
        output_path = os.path.join(output_folder, photo_dir_path)
        os.makedirs(output_path, exist_ok=True)

        for photo in photos_arr:
            files_iterator += 1
            if files_iterator > max_photos > 0:
                break

            link_format = os.path.splitext(photo)[1]
            if link_format:
                file_format = link_format
            else:
                file_format = default_file_format

            filename_art = prepared_art + '_' + str(files_iterator) + file_format

            r = http.request('GET', photo, preload_content=False)
            write_path = os.path.join(output_path, filename_art)
            with open(write_path, 'wb+') as out:
                while True:
                    data = r.read(data_chunks)  # default 2**16
                    if not data:
                        break
                    out.write(data)
            site_file_path = os.path.join(site_path, photo_dir_path, filename_art)
            clean_path = os.path.normpath(site_file_path)
            if nix_paths:
                clean_path = clean_path.replace('\\', '/')
            else:
                clean_path = clean_path.replace('/', '\\')
            output_files_struct[art].append(clean_path)
            r.release_conn()

    if empty_folders_counter:
        print(empty_folders_counter, 'skipped')
    return output_files_struct


if __name__ == '__main__':
    config = prepare_config('IMGdownloader.ini')
    if config:
        url_replace = str.split(config['global_settings']['url_replace'], ',')
        data_chunks = int(config['global_settings']['data_chunks'])
        max_photos = int(config['global_settings']['max_photos'])
        first_row = int(config['global_settings']['first_row'])
        last_row = int(config['global_settings']['last_row'])
        if last_row <= 0:
            last_row = None
        default_file_format = config['global_settings']['default_file_format']
        skip_trailing_slashes = config['global_settings'].getboolean('skip_trailing_slashes')

        create_art_folders = config['paths'].getboolean('create_art_folders')
        input_file = config['paths']['input_file']
        output_file = config['paths']['output_file']
        site_path = config['paths']['site_path']
        output_folder = config['paths']['output_folder']
        nix_paths = config['paths'].getboolean('nix_paths')

        file_struct = read_excel(input_file)
        links_struct = download_art_photo(file_struct)
        write_excel(links_struct)

        input("Press Enter to continue...")
